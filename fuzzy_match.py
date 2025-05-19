import pandas as pd
from rapidfuzz import process, fuzz
from tqdm import tqdm
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def normalize_name(name):
    name = str(name).lower()
    # Handle singkatan nama
    name = re.sub(r'\bm\.?\b', 'moch', name)
    name = re.sub(r'\bs\.?\b', 'siti', name)
    name = ' '.join([word for word in name.split() if len(word) > 1])
    name = re.sub(r'[^a-z0-9\s]', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

# Buat folder output jika belum ada
os.makedirs('output', exist_ok=True)

# Load data
data_lengkap = pd.read_excel('data_lengkap.xlsx', sheet_name=None)
data_email_benar = pd.read_excel('data_email_benar.xlsx', sheet_name=None, header=0)
sheet_names_lengkap = [f"XII-{i}" for i in range(1, 13)]
daftar_tidak_cocok = []
processed_sheets = {}

# Proses pencocokan untuk setiap sheet
for sheet in tqdm(sheet_names_lengkap, desc='Memproses semua kelas'):
    # Persiapan data
    df_lengkap = data_lengkap[sheet].copy()
    df_email = data_email_benar[sheet].copy()  # FIXED: gunakan nama sheet lengkap seperti "XII-1"
    
    # Normalisasi nama
    df_lengkap['nama_siswa_norm'] = df_lengkap['nama_siswa'].apply(normalize_name)
    df_email['NAMA_norm'] = df_email['NAMA'].apply(normalize_name)
    email_att_dict = dict(zip(df_email['NAMA_norm'], 
                           zip(df_email['EMAIL'], df_email['HADIR'])))
    
    # Insert kolom kehadiran
    df_lengkap.insert(3, 'kehadiran', '???')
    
    # Cek email asli
    has_email_column = 'email' in df_lengkap.columns
    original_emails = df_lengkap['email'].copy() if has_email_column else [''] * len(df_lengkap)
    
    # Proses pencocokan
    matched_emails = []
    attendances = []
    matched_flags = []
    for idx, name in enumerate(df_lengkap['nama_siswa_norm']):
        match = process.extractOne(name, email_att_dict.keys(), scorer=fuzz.partial_ratio)
        if match and match[1] > 80:
            email, attendance = email_att_dict[match[0]]
            matched_emails.append(email)
            attendances.append(attendance)
            matched_flags.append(True)
        else:
            matched_emails.append(original_emails.iloc[idx] if has_email_column else '')
            attendances.append("???")
            matched_flags.append(False)
            daftar_tidak_cocok.append({'kelas': sheet, 'nama': df_lengkap.iloc[idx]['nama_siswa']})
    
    # Update dataframe
    df_lengkap['email'] = matched_emails
    df_lengkap['kehadiran'] = attendances
    df_lengkap.drop(columns=['nama_siswa_norm'], inplace=True)
    processed_sheets[sheet] = (df_lengkap, matched_flags)

# Simpan ke file Excel dengan formatting
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

output_path = os.path.join('output', 'hasil_update_semua_kelas.xlsx')
daftar_tidak_path = os.path.join('output', 'daftar_tidak_cocok.xlsx')

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for sheet in sheet_names_lengkap:
        processed_sheets[sheet][0].to_excel(writer, sheet_name=sheet, index=False)

# Tambahkan formatting
wb = load_workbook(output_path)
for sheet in sheet_names_lengkap:
    ws = wb[sheet]
    flags = processed_sheets[sheet][1]
    for row_idx, flag in enumerate(flags, start=2):
        if not flag:
            for cell in ws[row_idx]:
                cell.fill = red_fill
wb.save(output_path)

# Simpan daftar tidak tercocok
pd.DataFrame(daftar_tidak_cocok).to_excel(daftar_tidak_path, index=False)

print(f'Proses selesai! File hasil disimpan di folder: {os.path.abspath("output")} tools: by nayyff and deepseek <3')
